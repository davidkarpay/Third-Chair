"""Parser for Axon Table_of_Contents.xlsx files."""

from pathlib import Path
from typing import Any, Optional

from ..models import Case, EvidenceItem


def find_toc_file(extracted_dir: Path) -> Optional[Path]:
    """
    Find the Table_of_Contents.xlsx file in extracted directory.

    Axon exports include a ToC file at the root level.

    Args:
        extracted_dir: Directory containing extracted files

    Returns:
        Path to ToC file if found, None otherwise
    """
    # Common ToC filenames
    toc_names = [
        "Table_of_Contents.xlsx",
        "TableOfContents.xlsx",
        "TOC.xlsx",
        "Contents.xlsx",
        "index.xlsx",
    ]

    # Check root level first
    for name in toc_names:
        toc_path = extracted_dir / name
        if toc_path.exists():
            return toc_path

    # Search recursively (max 2 levels)
    for name in toc_names:
        matches = list(extracted_dir.glob(f"*/{name}")) + list(extracted_dir.glob(f"**/{name}"))
        if matches:
            return matches[0]

    return None


def parse_toc(toc_path: Path) -> list[dict[str, Any]]:
    """
    Parse an Axon Table_of_Contents.xlsx file.

    The ToC typically contains:
    - Evidence ID
    - Filename
    - File type/category
    - Date/time
    - Officer information
    - Duration (for media)
    - Description

    Args:
        toc_path: Path to the ToC file

    Returns:
        List of dicts with evidence metadata
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required to parse Excel files. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(toc_path, read_only=True, data_only=True)
    ws = wb.active

    if ws is None:
        return []

    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(_normalize_header(cell.value) if cell.value else f"col_{len(headers)}")

    # Parse data rows
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        entry = {}
        for i, value in enumerate(row):
            if i < len(headers) and value is not None:
                entry[headers[i]] = value
        if entry:
            entries.append(entry)

    wb.close()
    return entries


def _normalize_header(header: str) -> str:
    """
    Normalize column header to consistent format.

    Converts various Axon column names to standardized keys.
    """
    if not header:
        return ""

    header = str(header).lower().strip()

    # Common column name mappings
    mappings = {
        "evidence id": "evidence_id",
        "evidenceid": "evidence_id",
        "id": "evidence_id",
        "file name": "filename",
        "filename": "filename",
        "file": "filename",
        "name": "filename",
        "file type": "file_type",
        "filetype": "file_type",
        "type": "file_type",
        "category": "category",
        "date": "date",
        "created": "date",
        "created date": "date",
        "time": "time",
        "created time": "time",
        "duration": "duration",
        "length": "duration",
        "officer": "officer",
        "officer name": "officer",
        "badge": "badge_number",
        "badge number": "badge_number",
        "badge #": "badge_number",
        "unit": "unit",
        "unit number": "unit",
        "description": "description",
        "notes": "notes",
        "comments": "notes",
    }

    # Remove special characters
    normalized = "".join(c if c.isalnum() or c.isspace() else "_" for c in header)
    normalized = "_".join(normalized.split())

    return mappings.get(header, normalized)


def update_case_from_toc(case: Case, toc_entries: list[dict[str, Any]]) -> Case:
    """
    Update case evidence items with metadata from ToC.

    Matches ToC entries to existing evidence items by filename
    and updates their metadata.

    Args:
        case: Case to update
        toc_entries: Parsed ToC entries

    Returns:
        Updated case
    """
    # Create filename lookup
    evidence_by_name: dict[str, EvidenceItem] = {}
    for evidence in case.evidence_items:
        evidence_by_name[evidence.filename.lower()] = evidence
        # Also index by stem (without extension)
        evidence_by_name[Path(evidence.filename).stem.lower()] = evidence

    # Match and update
    for entry in toc_entries:
        filename = entry.get("filename", "")
        if not filename:
            continue

        filename_lower = str(filename).lower()
        evidence = evidence_by_name.get(filename_lower)

        if not evidence:
            # Try matching by stem
            stem = Path(filename_lower).stem
            evidence = evidence_by_name.get(stem)

        if evidence:
            _update_evidence_from_toc(evidence, entry)

    return case


def _update_evidence_from_toc(evidence: EvidenceItem, toc_entry: dict[str, Any]) -> None:
    """Update an evidence item with ToC metadata."""
    # Update evidence ID if provided
    if toc_entry.get("evidence_id"):
        evidence.id = str(toc_entry["evidence_id"])

    # Parse duration if provided
    if toc_entry.get("duration"):
        duration = _parse_duration(toc_entry["duration"])
        if duration:
            evidence.duration_seconds = duration

    # Store other metadata
    metadata_keys = ["officer", "badge_number", "unit", "description", "notes", "category"]
    for key in metadata_keys:
        if toc_entry.get(key):
            evidence.metadata[key] = toc_entry[key]

    # Store date/time
    if toc_entry.get("date"):
        evidence.metadata["toc_date"] = str(toc_entry["date"])
    if toc_entry.get("time"):
        evidence.metadata["toc_time"] = str(toc_entry["time"])


def _parse_duration(duration_str: Any) -> Optional[float]:
    """
    Parse duration string to seconds.

    Handles formats like:
    - "1:23:45" (HH:MM:SS)
    - "5:30" (MM:SS)
    - "90" (seconds)
    - "1h 30m 45s"
    """
    import re

    if isinstance(duration_str, (int, float)):
        return float(duration_str)

    if not isinstance(duration_str, str):
        return None

    duration_str = duration_str.strip()

    # HH:MM:SS or MM:SS format
    time_match = re.match(r"^(\d+):(\d{2}):(\d{2})$", duration_str)
    if time_match:
        hours, minutes, seconds = map(int, time_match.groups())
        return hours * 3600 + minutes * 60 + seconds

    time_match = re.match(r"^(\d+):(\d{2})$", duration_str)
    if time_match:
        minutes, seconds = map(int, time_match.groups())
        return minutes * 60 + seconds

    # Plain number (seconds)
    try:
        return float(duration_str)
    except ValueError:
        pass

    # Natural language format
    total = 0.0
    hour_match = re.search(r"(\d+)\s*h", duration_str, re.IGNORECASE)
    if hour_match:
        total += int(hour_match.group(1)) * 3600

    min_match = re.search(r"(\d+)\s*m(?:in)?", duration_str, re.IGNORECASE)
    if min_match:
        total += int(min_match.group(1)) * 60

    sec_match = re.search(r"(\d+)\s*s(?:ec)?", duration_str, re.IGNORECASE)
    if sec_match:
        total += int(sec_match.group(1))

    return total if total > 0 else None
