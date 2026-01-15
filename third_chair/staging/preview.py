"""Quick preview of Axon ZIP files without full extraction."""

import io
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional


@dataclass
class ZipPreview:
    """Preview metadata for a staged Axon ZIP file."""

    zip_path: Path
    zip_size_mb: float
    case_id: str
    court_case: Optional[str] = None
    agency: Optional[str] = None
    file_count: int = 0
    file_types: dict[str, int] = field(default_factory=dict)
    has_toc: bool = False
    toc_entries: list[dict[str, Any]] = field(default_factory=list)
    validation_errors: list[str] = field(default_factory=list)

    @property
    def video_count(self) -> int:
        """Number of video files."""
        return self.file_types.get("video", 0)

    @property
    def audio_count(self) -> int:
        """Number of audio files."""
        return self.file_types.get("audio", 0)

    @property
    def document_count(self) -> int:
        """Number of document files."""
        return self.file_types.get("document", 0)

    @property
    def image_count(self) -> int:
        """Number of image files."""
        return self.file_types.get("image", 0)


# File extension to type mappings
FILE_TYPE_EXTENSIONS = {
    "video": {".mp4", ".mov", ".avi", ".mkv", ".wmv", ".m4v", ".webm"},
    "audio": {".mp3", ".wav", ".m4a", ".aac", ".ogg", ".wma", ".flac"},
    "document": {".pdf", ".doc", ".docx", ".txt", ".rtf", ".xlsx", ".xls", ".csv"},
    "image": {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff", ".tif", ".webp"},
}


def _classify_file_type(filename: str) -> str:
    """Classify a file by its extension."""
    ext = Path(filename).suffix.lower()
    for file_type, extensions in FILE_TYPE_EXTENSIONS.items():
        if ext in extensions:
            return file_type
    return "other"


def _extract_case_id_from_filename(filename: str) -> str:
    """Extract case ID from Axon ZIP filename."""
    patterns = [
        r"Case[_-]?(\d{4}[_-]\d+)",  # Case_2025-12345
        r"(\d{4}[_-]CF[_-]\d+)",  # 2025-CF-001234
        r"(\d{4}[_-]MM[_-]\d+)",  # 2025-MM-001234
        r"(\d{4}[_-]\d{5,})",  # 2025-12345
    ]

    for pattern in patterns:
        match = re.search(pattern, filename, re.IGNORECASE)
        if match:
            return match.group(1).replace("_", "-")

    # Fallback: use filename stem
    stem = Path(filename).stem
    return stem[:30] if len(stem) > 30 else stem


def _find_toc_in_zip(zf: zipfile.ZipFile) -> Optional[str]:
    """Find Table of Contents file path within ZIP."""
    toc_names = [
        "Table_of_Contents.xlsx",
        "TableOfContents.xlsx",
        "TOC.xlsx",
        "Contents.xlsx",
        "index.xlsx",
    ]

    for info in zf.infolist():
        filename = Path(info.filename).name
        if filename in toc_names:
            return info.filename

    return None


def _parse_toc_from_bytes(toc_bytes: bytes) -> list[dict[str, Any]]:
    """Parse ToC entries from xlsx bytes."""
    try:
        import openpyxl
    except ImportError:
        return []

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(toc_bytes),
            read_only=True,
            data_only=True
        )
        ws = wb.active

        if ws is None:
            return []

        # Get headers from first row
        headers = []
        for cell in ws[1]:
            header = _normalize_header(cell.value) if cell.value else f"col_{len(headers)}"
            headers.append(header)

        # Parse data rows
        entries = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            entry = {}
            for i, value in enumerate(row):
                if i < len(headers) and value is not None:
                    entry[headers[i]] = value
            if entry:
                entries.append(entry)

        wb.close()
        return entries

    except Exception:
        return []


def _normalize_header(header: Any) -> str:
    """Normalize column header to consistent format."""
    if not header:
        return ""

    header = str(header).lower().strip()

    mappings = {
        "evidence id": "evidence_id",
        "evidenceid": "evidence_id",
        "id": "evidence_id",
        "file name": "filename",
        "filename": "filename",
        "file": "filename",
        "name": "filename",
        "file type": "file_type",
        "filetype": "file_type",
        "type": "file_type",
        "category": "category",
        "date": "date",
        "created": "date",
        "created date": "date",
        "time": "time",
        "created time": "time",
        "duration": "duration",
        "length": "duration",
        "officer": "officer",
        "officer name": "officer",
        "badge": "badge_number",
        "badge number": "badge_number",
        "badge #": "badge_number",
        "unit": "unit",
        "unit number": "unit",
        "description": "description",
        "notes": "notes",
        "comments": "notes",
    }

    normalized = "".join(c if c.isalnum() or c.isspace() else "_" for c in header)
    normalized = "_".join(normalized.split())

    return mappings.get(header, normalized)


def _detect_agency(toc_entries: list[dict], zip_filename: str) -> Optional[str]:
    """Try to detect agency from ToC entries or filename."""
    # Check ToC entries for agency field
    for entry in toc_entries:
        agency = entry.get("agency") or entry.get("department")
        if agency:
            return str(agency)

    # Common agency patterns in filenames
    agency_patterns = [
        (r"PBSO", "Palm Beach County Sheriff's Office"),
        (r"BSO", "Broward Sheriff's Office"),
        (r"MDPD", "Miami-Dade Police Department"),
        (r"FHP", "Florida Highway Patrol"),
        (r"FDLE", "Florida Department of Law Enforcement"),
    ]

    for pattern, name in agency_patterns:
        if re.search(pattern, zip_filename, re.IGNORECASE):
            return name

    return None


def preview_axon_zip(zip_path: Path) -> ZipPreview:
    """
    Generate a quick preview of an Axon ZIP file without full extraction.

    Args:
        zip_path: Path to the ZIP file.

    Returns:
        ZipPreview with metadata and ToC entries.
    """
    errors = []

    if not zip_path.exists():
        return ZipPreview(
            zip_path=zip_path,
            zip_size_mb=0.0,
            case_id="unknown",
            validation_errors=[f"File not found: {zip_path}"],
        )

    if not zipfile.is_zipfile(zip_path):
        return ZipPreview(
            zip_path=zip_path,
            zip_size_mb=zip_path.stat().st_size / (1024 * 1024),
            case_id="unknown",
            validation_errors=[f"Not a valid ZIP file: {zip_path}"],
        )

    # Get file size
    zip_size_mb = zip_path.stat().st_size / (1024 * 1024)

    # Extract case ID from filename
    case_id = _extract_case_id_from_filename(zip_path.name)

    # Analyze ZIP contents
    file_types: dict[str, int] = {}
    file_count = 0

    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            # Count files by type
            for info in zf.infolist():
                if info.is_dir() or info.filename.startswith("__MACOSX"):
                    continue
                if Path(info.filename).name.startswith("."):
                    continue

                file_count += 1
                file_type = _classify_file_type(info.filename)
                file_types[file_type] = file_types.get(file_type, 0) + 1

            # Try to read ToC
            toc_path = _find_toc_in_zip(zf)
            has_toc = toc_path is not None
            toc_entries = []

            if toc_path:
                try:
                    toc_bytes = zf.read(toc_path)
                    toc_entries = _parse_toc_from_bytes(toc_bytes)
                except Exception as e:
                    errors.append(f"Could not read ToC: {e}")

    except zipfile.BadZipFile as e:
        errors.append(f"Corrupt ZIP file: {e}")
        return ZipPreview(
            zip_path=zip_path,
            zip_size_mb=zip_size_mb,
            case_id=case_id,
            validation_errors=errors,
        )

    # Detect court case number from ToC
    court_case = None
    for entry in toc_entries:
        if "case" in str(entry).lower():
            # Look for case number patterns
            for value in entry.values():
                if value and re.search(r"\d{4}-?[A-Z]{2}-?\d+", str(value)):
                    court_case = str(value)
                    break
        if court_case:
            break

    # Detect agency
    agency = _detect_agency(toc_entries, zip_path.name)

    return ZipPreview(
        zip_path=zip_path,
        zip_size_mb=zip_size_mb,
        case_id=case_id,
        court_case=court_case,
        agency=agency,
        file_count=file_count,
        file_types=file_types,
        has_toc=has_toc,
        toc_entries=toc_entries,
        validation_errors=errors,
    )
